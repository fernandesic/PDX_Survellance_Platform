"""
Parse the Ebola/DRC sitrep tracker xlsx uploaded by superadmin and write
its rows into the OutbreakEvent stream.

Source of truth for case counts on the outbreak page. Three sheets matter:

  - `manual_sitrepCumulatives` — one row per sitrep date. Country-level
    cumulative + delta counts. Drives the headline tile strip.
  - `daily_cumul_counts`       — Province × Health Zone × Date rows.
    Drives the by-zone panel and the map.
  - `daily_testing`            — daily lab throughput. Drives the lab panel.

All idempotent on (source, source_ref). Reuploading the same file is a
no-op. New rows for new dates land as new events.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone as dt_timezone
from typing import Any, IO

import openpyxl
from django.db import IntegrityError

from outbreak.models import EventKind, Outbreak, OutbreakEvent

logger = logging.getLogger(__name__)


TRACKER_SOURCE = 'manual_tracker'

# Sheet names exactly as they appear in the workbook.
SHEET_DAILY = 'manual_sitrepCumulatives'
SHEET_ZONE = 'daily_cumul_counts'
SHEET_LAB = 'daily_testing'

# Column headers we expect — looked up case-insensitively. We accept the
# sheet's mild label drift ("Confirmed Deaths" vs "Confirmed_Deaths") by
# normalising on lookup.
DAILY_COLUMNS = {
    'date': ['date'],
    'new_suspected_cases': ['newsuspected_cases', 'new suspected cases'],
    'suspected_cases': ['suspected_cases', 'suspected cases'],
    'suspected_deaths': ['suspected_deaths', 'suspected deaths'],
    'new_confirmed_cases': ['newconfirmed_cases', 'new confirmed cases'],
    'confirmed_cases': ['confirmed_cases', 'confirmed cases'],
    'new_confirmed_deaths': ['newconfirmed_deaths', 'new confirmed deaths'],
    'confirmed_deaths': ['confirmed_deaths', 'confirmed deaths'],
    'new_contacts': ['new_contacts', 'new contacts'],
    'total_contacts': ['total_contacts', 'total contacts'],
    'verification': [''],  # the 14th column in the sheet has no header
}

ZONE_COLUMNS = {
    'as_of': ['reporting date', 'sitrep date'],
    'country': ['country'],
    'province': ['province'],
    'health_zone': ['health zone'],
    'suspected_cases': ['suspected cases'],
    'suspected_deaths': ['suspected deaths'],
    'confirmed_cases': ['confirmed cases'],
    'confirmed_deaths': ['confirmed deaths'],
    'new_suspected_cases': ['new suspected cases'],
    'new_suspected_deaths': ['new suspected deaths'],
    'new_confirmed_cases': ['new confirmed cases'],
    'new_confirmed_deaths': ['new confirmed deaths'],
    'contacts': ['contacts'],
    'clean_province': ['clean province'],
    'clean_health_zone': ['clean health zone'],
    'source': ['source'],
}

LAB_COLUMNS = {
    'as_of': ['date'],
    'country': ['country'],
    'province': ['province'],
    'health_zone': ['health zone'],
    'samples_received': ['samples received'],
    'samples_analyzed': ['samples analyzed'],
    'positive_samples': ['positive samples'],
    'test_positivity': ['test positivity'],
    'daily_throughput': ['daily throughput'],
    'backlog': ['backlog of samples', 'backlog'],
}


class TrackerParseError(Exception):
    pass


def ingest_tracker(upload: IO[bytes], outbreak: Outbreak) -> dict:
    """
    Parse the xlsx file-like object and write events to `outbreak`.
    Returns a summary dict for the UI to display after upload.
    """
    try:
        wb = openpyxl.load_workbook(upload, data_only=True, read_only=True)
    except (OSError, ValueError, KeyError, openpyxl.utils.exceptions.InvalidFileException) as e:
        raise TrackerParseError(f'Could not open workbook: {e}') from e

    summary = {
        'daily_emitted': 0, 'daily_skipped': 0, 'daily_total': 0,
        'zone_emitted': 0,  'zone_skipped': 0,  'zone_total': 0,
        'lab_emitted': 0,   'lab_skipped': 0,   'lab_total': 0,
        'latest_as_of': None,
        'warnings': [],
    }

    if SHEET_DAILY in wb.sheetnames:
        d = _ingest_daily(wb[SHEET_DAILY], outbreak)
        summary['daily_emitted'] = d['emitted']
        summary['daily_skipped'] = d['skipped']
        summary['daily_total'] = d['total']
        summary['latest_as_of'] = d['latest_as_of']
    else:
        summary['warnings'].append(f"Sheet '{SHEET_DAILY}' missing")

    if SHEET_ZONE in wb.sheetnames:
        d = _ingest_zone(wb[SHEET_ZONE], outbreak)
        summary['zone_emitted'] = d['emitted']
        summary['zone_skipped'] = d['skipped']
        summary['zone_total'] = d['total']
    else:
        summary['warnings'].append(f"Sheet '{SHEET_ZONE}' missing")

    if SHEET_LAB in wb.sheetnames:
        d = _ingest_lab(wb[SHEET_LAB], outbreak)
        summary['lab_emitted'] = d['emitted']
        summary['lab_skipped'] = d['skipped']
        summary['lab_total'] = d['total']
    # Lab sheet is optional — no warning if absent.

    wb.close()
    return summary


# ── sheet ingesters ──────────────────────────────────────────────────────


def _ingest_daily(ws, outbreak: Outbreak) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _empty()
    idx = _header_index(rows[0], DAILY_COLUMNS)
    emitted = skipped = total = 0
    latest_as_of: str | None = None

    for raw in rows[1:]:
        as_of_dt = _as_date(raw[idx['date']]) if 'date' in idx else None
        if not as_of_dt:
            continue
        as_of_iso = as_of_dt.date().isoformat()
        total += 1

        payload = {
            'as_of': as_of_iso,
            'suspected_cases':       raw[idx['suspected_cases']]       if 'suspected_cases'       in idx else None,
            'suspected_deaths':      raw[idx['suspected_deaths']]      if 'suspected_deaths'      in idx else None,
            'confirmed_cases':       raw[idx['confirmed_cases']]       if 'confirmed_cases'       in idx else None,
            'confirmed_deaths':      raw[idx['confirmed_deaths']]      if 'confirmed_deaths'      in idx else None,
            'new_suspected_cases':   raw[idx['new_suspected_cases']]   if 'new_suspected_cases'   in idx else None,
            'new_confirmed_cases':   raw[idx['new_confirmed_cases']]   if 'new_confirmed_cases'   in idx else None,
            'new_confirmed_deaths':  raw[idx['new_confirmed_deaths']]  if 'new_confirmed_deaths'  in idx else None,
            'new_contacts':          raw[idx['new_contacts']]          if 'new_contacts'          in idx else None,
            'total_contacts':        raw[idx['total_contacts']]        if 'total_contacts'        in idx else None,
            'verification':          raw[idx['verification']]          if 'verification'          in idx else None,
        }
        if _all_blank(payload, except_keys={'as_of'}):
            continue

        source_ref = f"manual_tracker:daily:{outbreak.iso3.upper()}:{as_of_iso}"
        if _emit(outbreak, as_of_dt, EventKind.CASE, '', payload, source_ref):
            emitted += 1
            if not latest_as_of or as_of_iso > latest_as_of:
                latest_as_of = as_of_iso
        else:
            skipped += 1

    return {
        'emitted': emitted, 'skipped': skipped, 'total': total,
        'latest_as_of': latest_as_of,
    }


def _ingest_zone(ws, outbreak: Outbreak) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _empty()
    idx = _header_index(rows[0], ZONE_COLUMNS)
    emitted = skipped = total = 0

    for raw in rows[1:]:
        as_of_dt = _as_date(raw[idx['as_of']]) if 'as_of' in idx else None
        if not as_of_dt:
            continue
        province = _str(raw[idx['province']]) if 'province' in idx else ''
        zone     = _str(raw[idx['health_zone']]) if 'health_zone' in idx else ''
        if not province and not zone:
            continue
        total += 1
        as_of_iso = as_of_dt.date().isoformat()

        payload = {
            'as_of': as_of_iso,
            'country':              _str(raw[idx['country']]) if 'country' in idx else '',
            'province':             province,
            'health_zone':          zone,
            'suspected_cases':      raw[idx['suspected_cases']]      if 'suspected_cases'      in idx else None,
            'suspected_deaths':     raw[idx['suspected_deaths']]     if 'suspected_deaths'     in idx else None,
            'confirmed_cases':      raw[idx['confirmed_cases']]      if 'confirmed_cases'      in idx else None,
            'confirmed_deaths':     raw[idx['confirmed_deaths']]     if 'confirmed_deaths'     in idx else None,
            'new_suspected_cases':  raw[idx['new_suspected_cases']]  if 'new_suspected_cases'  in idx else None,
            'new_confirmed_cases':  raw[idx['new_confirmed_cases']]  if 'new_confirmed_cases'  in idx else None,
            'contacts':             raw[idx['contacts']]             if 'contacts'             in idx else None,
            'source':               _str(raw[idx['source']]) if 'source' in idx else '',
        }
        if _all_blank(payload, except_keys={'as_of', 'country', 'province', 'health_zone', 'source'}):
            continue

        # Dedup key includes date so reuploads of the same sheet are no-ops
        # but new daily entries for the same zone do land.
        key = f"{province}|{zone}|{as_of_iso}".lower()
        source_ref = f"manual_tracker:zone:{outbreak.iso3.upper()}:{key}"
        if _emit(outbreak, as_of_dt, EventKind.CASE, f'{province}/{zone}'.strip('/'), payload, source_ref):
            emitted += 1
        else:
            skipped += 1
    return {'emitted': emitted, 'skipped': skipped, 'total': total}


def _ingest_lab(ws, outbreak: Outbreak) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _empty()
    idx = _header_index(rows[0], LAB_COLUMNS)
    emitted = skipped = total = 0

    for raw in rows[1:]:
        as_of_dt = _as_date(raw[idx['as_of']]) if 'as_of' in idx else None
        if not as_of_dt:
            continue
        total += 1
        as_of_iso = as_of_dt.date().isoformat()
        payload = {
            'as_of': as_of_iso,
            'country':           _str(raw[idx['country']]) if 'country' in idx else '',
            'province':          _str(raw[idx['province']]) if 'province' in idx else '',
            'health_zone':       _str(raw[idx['health_zone']]) if 'health_zone' in idx else '',
            'samples_received':  raw[idx['samples_received']]  if 'samples_received'  in idx else None,
            'samples_analyzed':  raw[idx['samples_analyzed']]  if 'samples_analyzed'  in idx else None,
            'positive_samples':  raw[idx['positive_samples']]  if 'positive_samples'  in idx else None,
            'test_positivity':   raw[idx['test_positivity']]   if 'test_positivity'   in idx else None,
            'daily_throughput':  raw[idx['daily_throughput']]  if 'daily_throughput'  in idx else None,
            'backlog':           raw[idx['backlog']]           if 'backlog'           in idx else None,
        }
        if _all_blank(payload, except_keys={'as_of', 'country', 'province', 'health_zone'}):
            continue

        source_ref = f"manual_tracker:lab:{outbreak.iso3.upper()}:{as_of_iso}"
        # 'lab' isn't in EventKind — use 'case' kind so adaptor base
        # validators don't warn. The source_ref prefix distinguishes lab
        # rows for the snapshot service.
        if _emit(outbreak, as_of_dt, EventKind.CASE, '', payload, source_ref):
            emitted += 1
        else:
            skipped += 1
    return {'emitted': emitted, 'skipped': skipped, 'total': total}


# ── helpers ──────────────────────────────────────────────────────────────


def _header_index(header_row: tuple, expected: dict[str, list[str]]) -> dict[str, int]:
    """Build a {logical-key → column-index} map, tolerant to label drift."""
    headers = [_norm_header(h) for h in header_row]
    idx: dict[str, int] = {}
    for key, candidates in expected.items():
        for cand in candidates:
            if cand in headers:
                idx[key] = headers.index(cand)
                break
        else:
            # Header missing for this key — log once, continue.
            logger.debug(
                'tracker_xlsx: header %r not found in row (have: %s)',
                candidates, headers,
            )
    return idx


def _norm_header(h: Any) -> str:
    if h is None:
        return ''
    return str(h).strip().lower().replace('\xa0', ' ').replace('  ', ' ')


def _as_date(v) -> datetime | None:
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        dt = v
    else:
        try:
            dt = datetime.fromisoformat(str(v).split(' ')[0])
        except (TypeError, ValueError):
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=dt_timezone.utc)
    return dt


def _str(v) -> str:
    if v is None:
        return ''
    return str(v).strip().replace('\xa0', ' ').replace('  ', ' ')


def _all_blank(payload: dict, except_keys: set[str]) -> bool:
    for k, v in payload.items():
        if k in except_keys:
            continue
        if v is None or v == '' or v == 0:
            continue
        return False
    return True


def _emit(outbreak: Outbreak, ts: datetime, kind: str, geo: str,
          payload: dict, source_ref: str) -> bool:
    """Insert one OutbreakEvent. Returns True if new, False if dedup-skipped."""
    # Strip Nones / blanks the JSON doesn't need.
    payload = {k: v for k, v in payload.items() if v not in (None, '')}
    if OutbreakEvent.objects.filter(
        source=TRACKER_SOURCE, source_ref=source_ref,
    ).exists():
        return False
    try:
        OutbreakEvent.objects.create(
            outbreak=outbreak,
            ts=ts,
            source=TRACKER_SOURCE,
            kind=kind,
            geo=geo,
            payload_json=payload,
            confidence=1.00,
            source_ref=source_ref,
        )
        return True
    except IntegrityError:
        return False


def _empty() -> dict:
    return {'emitted': 0, 'skipped': 0, 'total': 0, 'latest_as_of': None}
